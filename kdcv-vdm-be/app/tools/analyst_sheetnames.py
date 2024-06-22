import openpyxl
import os

file_name = "T24Y003V00-VDM001-AB.xlsb"
folder_path = "/home/amrserver/Documents/KDCV-m104/xlsx"

# workbook = openpyxl.load_workbook(os.path.join(folder_path, file_name))

def find_all_xlsx_in_folder(folder_path):
    listFile = []
    for filename in os.listdir(folder_path):
        # Kiểm tra xem file có phần mở rộng .xlsx hay không
        if filename.endswith(('.xlsx','.XLSX')):
            # Tạo đường dẫn đầy đủ tới file .xlsx
            xlsx_file = os.path.join(folder_path, filename)
            listFile.append(xlsx_file)
    return listFile

allFileXLSX = find_all_xlsx_in_folder(folder_path)

i = 0
sheetNameNhanvien = []
for file in allFileXLSX:
    print("//////////////////////////////////////")
    print("name file: ", file)
    workbook = openpyxl.load_workbook(file)
    sheetNames = workbook.sheetnames
    print(f"Sheet names: {sheetNames}")
    flag = False
    for name in sheetNames:
        if name.startswith('N'):
            if not flag:
                i += 1
                flag = True
            if name not in sheetNameNhanvien:
                sheetNameNhanvien.append(name)
            
print("Tổng số file có sheetname bắt đầu bằng N: ", i)
print("Tổng hợp tên sheet giành cho Nhân viên: ", sheetNameNhanvien)

# print(sheet.columns)
# not_nan_pos = sheet.notna().stack()
# print("Các giá trị không phải NaN cùng với vị trí hàng và cột:")
# for position, value in not_nan_pos.items():
#     if value:
#         row, column = position
#         print(f"Giá trị: {sheet.iloc[row][column]}, Vị trí hàng: {row}, Vị trí cột: {column}")

# for idx, value in enumerate(sheet[3]):
#     print(f"Hàng số {idx + 1}: {value}")

# col = 1
# not_nan_row_pos = sheet[col].dropna().index

# print(not_nan_row_pos)
# for i in not_nan_row_pos:
#     print(f"value at index row-{i}: {sheet.at[i,col]}, type: {type(sheet.at[i,col])}")