# from openpyxl import load_workbook
# from openpyxl_image_loader import SheetImageLoader
# import os


# file_name = "T24Y003V00-VDM001-AB.xlsx"
# folder_path = "/home/amrserver/Documents/KDCV-m104/xlsx"

# def find_all_xlsx_in_folder(folder_path):
#     listFile = []
#     for filename in os.listdir(folder_path):
#         # Kiểm tra xem file có phần mở rộng .xlsb hay không
#         if filename.endswith(('.xlsx','.XLSX')):
#             # Tạo đường dẫn đầy đủ tới file .xlsb
#             xlsb_file = os.path.join(folder_path, filename)
#             listFile.append(xlsb_file)
#     return listFile

# # Load your workbook and sheet as you want, for example
# wb = load_workbook(os.path.join(folder_path, file_name))
# sheet = wb['KTV (2)']

# allFileXLSX = find_all_xlsx_in_folder(folder_path)
# # for file in allFileXLSX:
# #     wb = load_workbook(file)
# #     wb.close()


# # Put your sheet in the loader
# image_loader = SheetImageLoader(sheet)


# # # # And get image from specified cell
# image = image_loader.get('E5')

# print(image.info)
# # # # Image now is a Pillow image, so you can do the following
# image.show()

# # # Ask if there's an image in a cell
# # if image_loader.image_in('E5'):
# #     print("Got it!")
# # else:
# #     print("image not in cell you choose!")




from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Đường dẫn đến file Excel của bạn
file_path = '/home/amrserver/Documents/KDCV-m104/xlsx/T24Y003V00-VDM001-AB.xlsx'

# Mở workbook
wb = load_workbook(filename=file_path)

# Chọn sheet bạn muốn đọc
ws = wb['KTV (2)']  # Thay 'Sheet1' bằng tên sheet của bạn

# Danh sách các hình ảnh
images = []

# Duyệt qua tất cả các hình ảnh trong sheet
for image in ws._images:
    # Lấy tọa độ của hình ảnh
    anchor = image.anchor._from
    # Chuyển tọa độ thành địa chỉ ô
    cell_address = ws.cell(anchor.row + 1, anchor.col + 1).coordinate
    images.append((cell_address, image))

# Duyệt qua các hình ảnh và in thông tin
for idx, (cell_address, image) in enumerate(images):
    print(f'Hình ảnh tại ô {cell_address}')
    # Nếu bạn muốn lưu hình ảnh
    image_path = f'image_from_{cell_address}_{idx}.png'
    with open(image_path, 'wb') as img_file:
        img_file.write(image._data())
    print(f'Hình ảnh từ ô {cell_address} đã được lưu tại {image_path}')

# Đóng workbook
wb.close()
