import os
import re
import openpyxl

# file_name = "TDB079BV00-VDM001-AD.xlsx"
folder_path = "/home/amrserver/Documents/KDCV-m104/xlsx"

# sheet = pd.read_excel(folder_path+file_name, engine="pyxlsb", sheet_name="NVTT 1", usecols='A:J' , header=None)

class StructForm:
    def __init__(self) -> None:
        self.operationName = "A1"
        self.year = [1,1]
        self.speciesName = [1,3]
        self.month = [2,1]
        self.shiftStage = [2,3]
        self.deviceName = "A5"
        self.no = [4,2]
        self.category = [3,2]
        self.method = [3,3]
        self.position = [3,4]
        self.confirm = [3,5]
        self.type = [3,6]

def find_all_xlsx_in_folder(folder_path):
    listFile = []
    for filename in os.listdir(folder_path):
        # Kiểm tra xem file có phần mở rộng .xlsb hay không
        if filename.endswith(('.xlsx','.XLSX')):
            # Tạo đường dẫn đầy đủ tới file .xlsb
            # xlsb_file = os.path.join(folder_path, filename)
            xlsb_file = filename
            listFile.append(xlsb_file)
    return listFile

allFileXLSX = find_all_xlsx_in_folder(folder_path)

# Verify if files are supported
def is_excel_file(files):
    supported_files = ['.xls', '.xlsx', '.xlsb', '.XLSB']

    for file in files:
        file_type = os.path.splitext(file)[1]
        if file_type not in supported_files:
            return False
    
    return True

for excelFile in allFileXLSX:
    column_index = 2
    sheetNameNhanvien = []
    workbook = openpyxl.load_workbook(filename=os.path.join(folder_path, excelFile), data_only=True)
    sheetNames = workbook.sheetnames
    analyst_categories = []
    jobName = ""
    print(f"Mã số KDCV: {excelFile[:-5]}")
    for name in sheetNames:
        if name.startswith('K'):
            # print(f"Sheet name: {name} -----------------/")
            sheet = workbook[name]
            jobName = sheet['A1'].value
            # Lấy tất cả giá trị trong cột và vị trí của chúng
            non_nan_values = []
            # column_values = [cell.value for cell in sheet['H']]
            # print(f"Phân khu column({len(column_values)}): {column_values}")

            # for row in range(5, sheet.max_row + 1):
            #     cell_value = sheet.cell(row=row, column=column_index).value
            #     if (cell_value is not None and
            #         isinstance(cell_value, int)):
            #         print("///////////////////////////////////////////////")
            #         print(f"Hạng mục điểm kiểm {cell_value}: {sheet.cell(row=row, column=3).value}")
            #         print(f"  Phương pháp điểm kiểm: {sheet.cell(row=row, column=4).value}")
            #         print(f"  Phương pháp xác nhận & bản TMKT: {sheet.cell(row=row, column=6).value}")

            # Tìm tất cả các ô merged và gán chúng dựa vào key là tên cột start
            merged_cells = sheet.merged_cells.ranges
            merged_cells_filter = {}
            for merged_cell in merged_cells:
                addr = merged_cell.coord
                match = re.match(r"([A-Za-z]+)\d+:[A-Za-z]+\d+", addr).group(1)
                if match not in merged_cells_filter:
                    merged_cells_filter[match] = [merged_cell.coord]
                else:
                    merged_cells_filter[match].append(merged_cell.coord)
            # print(merged_cells_filter)
            
            # Tìm tất cả các ô có giá trị khác none và vị trí hàng của chúng trong cột bất kỳ
            for row in range(5, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=column_index).value
                if (cell_value is not None):
                    non_nan_values.append((cell_value, row))

            # Tìm vị trí hàng của từng hạng mục điểm kiểm
            for value, row in non_nan_values:
                if isinstance(value, int):
                    # print(f"Vị trí: Hàng {row}, Cột {column_index} - Giá trị: {value}")
                    category_merged = merged_cells_filter["C"]
                    # subdivisions = []
                    # for cell in category_merged:
                    #     start = int(cell.split(':')[0][1:])
                    #     end = int(cell.split(':')[1][1:])
                    #     if start == row:
                    #         isMeasure = False
                    #         for i in range(start, end + 1):
                    #             subName = sheet.cell(row=i, column=8).value
                    #             if (subName is not None):
                    #                 sub = {"name": subName}
                    #                 if (sheet.cell(row=i, column=9).value and
                    #                     sheet.cell(row=i, column=10).value):
                    #                     isMeasure = True
                    #                     rawDvd1 = sheet.cell(row=i, column=10).value
                    #                     rawDvd2 = sheet.cell(row=i+1, column=10).value
                    #                     dvd1 = rawDvd1.replace("\n", "").replace(" ","")
                    #                     dvd2 = rawDvd2.replace("\n", "").replace(" ","")
                    #                     sub["isMeasure"] = isMeasure
                    #                     sub["unitOfMeasurement"] = [dvd1, dvd2]
                                        
                    #                 else:
                    #                     sub["isMeasure"] = isMeasure
                    #                     sub["unitOfMeasurement"] = []
                    #                 subdivisions.append(sub)

                    # category[0]: step
                    # category[1]: hạng mục điểm kiểm
                    # category[2]: Phương pháp điểm kiểm
                    # category[3]: Vị trí điểm kiểm
                    # category[4]: Phương pháp xác nhận & bản TMKT
                    # category[5]: Chủng loại
                    # category[6]: Phân khu: [{name: tên phân khu, is_measure: trị thực đo, unitOfMeasurement: đơn vị đo}]
                    # category[7]: Các thay đổi cần ghi chép lại: {deviceChange: thay đổi thiết bị, partCodeChange: thay đổi mã hàng, powerFailure: sự cố về nguồn điện}
                    category = [0,0,0,0,0,0,0,0]
                    pattern_IVXLCDM = r"\([IVXLCDM ,.;:]+\)"
                    category[0] = value
                    category[1] = sheet.cell(row=row, column=3).value
                    # print(f"{category[0]}: {category[1]}")
                    if category[1] is not None:
                        category[1] = category[1].replace("\n"," ")
                        match_IVXLCDM = re.search(pattern_IVXLCDM, category[1])
                    else:
                        continue
                    
                    if match_IVXLCDM:
                        extracted_string = match_IVXLCDM.group()
                        normalized_string = re.sub(r"[ ,.;:]+", ',', extracted_string.strip('()'))
                        array_of_values = normalized_string.split(',')

                    # In kết quả
                    # print(f"Chuỗi đầu vào: '{category[1]}'")
                    # print(f"  Phần trích xuất: {array_of_values}")

                    category[2] = sheet.cell(row=row, column=4).value
                    if category[2] is not None:
                        category[2] = category[2].replace("\n"," ")
                    else:
                        continue

                    category[3] = sheet.cell(row=row, column=5).value

                    category[4] = sheet.cell(row=row, column=6).value
                    if category[4] is not None:
                        category[4] = category[4].replace("\n",", ")
                    else:
                        continue

                    category[5] = sheet.cell(row=row, column=7).value
                    if category[5] is not None:
                        category[5] = category[5].replace("\n",", ")

                    # category[6] = subdivisions
                    category[6] = {"deviceChange": False, "partCodeChange": False, "powerFailure": False}
                    for i in array_of_values:
                        if i == "I":
                            category[6]["deviceChange"] = True
                        elif i == "II":
                            category[6]["partCodeChange"] = True
                        elif i == "III":
                            category[6]["powerFailure"] = True

                    analyst_categories.append(category)

    patternJobName = r"tên\s*thao\s*tác\s*:\s*(.*)"
    matchJobName = re.search(patternJobName, jobName, re.IGNORECASE)
    if matchJobName:
        result = matchJobName.group(1)
        jobName = result
    else:
        print("Không tìm thấy 'Tên thao tác:' trong chuỗi đầu vào.")

    print(f"Tên thao tác: {jobName}")
    for category in analyst_categories:
        print(f"STT: {category[0]}")
        print(f"  Hạng mục điểm kiểm: {category[1]}")
        print(f"  Phương pháp điểm kiểm: {category[2]}") 
        print(f"  Vị trí điểm kiểm: {category[3]}") 
        print(f"  PP xác nhận & bản TMKT: {category[4]}")
        print(f"  Chủng loại: {category[5]}") 
        print(f"  Các thay đổi cần ghi chép lại: {category[6]}")
        # print(f"  Phân khu:")
        # for sub in category[6]:
        #     print(f"    Name: {sub['name']}, Trị thực đo: {sub['isMeasure']}, Đơn vị đo: {sub['unitOfMeasurement']}")
    print("/-------------------------------------------------------------/")

    # Đóng workbook
    workbook.close()

def extract_and_normalize(input_string):
    # Biểu thức chính quy để tìm các phần có dạng (I), (II), (I;III), (I.III), v.v.
    pattern = r"\([IVXLCDM ,.;:]+\)"
    
    # Tìm mẫu đầu tiên khớp với biểu thức chính quy
    match = re.search(pattern, input_string)
    if match:
        extracted_string = match.group()
        print(f"Phần trích xuất: {extracted_string}")
        
        # Chuẩn hóa các ký tự ngăn cách thành dấu phẩy
        normalized_string = re.sub(r"[ ,.;:]+", ',', extracted_string.strip('()'))
        print(f"Chuỗi chuẩn hóa: {normalized_string}")
        
        # Tách chuỗi thành mảng
        array_of_values = normalized_string.split(',')
        print(f"Mảng phân tích: {array_of_values}")
        
        return array_of_values
    else:
        print("Không tìm thấy phần khớp nào.")
        return []